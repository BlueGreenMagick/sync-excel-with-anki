import os

from .errors import *
from openpyxl import load_workbook, Workbook


class ExcelFileReadOnly:
    def __init__(self, path):
        self.path = path

    def load_file(self):
        self.wb = load_workbook(filename=self.path, read_only=True, data_only=True)
        ws = self.wb.worksheets[0]
        self.ws = ws
        self.wsv = ws.values

    def read_file(self):
        ws = self.ws
        wsv = self.wsv
        models = []
        models_fields = []
        models_desg = []
        log = ""

        # Get name of models in first row
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    model = str(cell.value).strip()
                    if model:
                        models.append(model)

        # Get name of fields per model
        for n in range(len(models)):
            for row in ws.iter_rows(min_row=n + 2, max_row=n + 2):
                model_fields = []
                for cell in row[1:]:
                    if cell.value and str(cell.value).strip():
                        cell = str(cell.value).strip()
                        model_fields.append(cell)
                models_fields.append(model_fields)
                models_desg.append(str(row[0].value).strip())
        self.models_fields = models_fields
        self.models = models
        self.models_desg = models_desg
        rows_data = []

        # Go through each note rows
        for row in ws.iter_rows(min_row=2 + len(models)):
            model_desg = row[0].value
            if not model_desg:
                log += ""
                continue
            model_desg = str(model_desg).strip()
            try:
                model_index = models_desg.index(model_desg)
            except Exception as e:
                self.close()
                raise InvalidModelDesignatorError(
                    self.path, row[0].row, model_desg
                ) from e
            model_name = models[model_index]
            model_fields = models_fields[model_index]
            nid = row[len(model_fields) + 1].value
            if nid:
                try:
                    nid = int(nid)
                except ValueError:
                    log += (
                        "\n<b>Non-fatal</b>: non integer value '%s' in nid field, in %d row"
                        % (str(nid), row[0].row)
                    )
                    nid = None
            else:
                nid = None
            row_data = {
                "row": row[0].row,
                "id": nid,
                "model": model_name,
                "fields": {},
                "log": log,
                "path": self.path,
            }  # row is 1 based
            # Get field values
            for i in range(0, len(model_fields)):
                if row[i + 1].value:
                    row_data["fields"][model_fields[i]] = str(row[i + 1].value)
                else:
                    row_data["fields"][model_fields[i]] = None
            rows_data.append(row_data)
        # [{"row":int, "id":int, "fields":{"fieldName":str_val,}, "model": str_model_name, "log": str_log}]
        return rows_data

    def close(self):
        self.wb.close()


class ExcelFile(ExcelFileReadOnly):
    def load_file(self):
        self.wb = load_workbook(filename=self.path)
        ws = self.wb.worksheets[0]
        self.ws = ws
        self.wsv = ws.values

    def set_id(self, row, fields, id):
        self.write_cell(row, len(fields) + 2, id)
        # cell = self.ws["A" + str(row_num)]
        # cell.value = id

    def create_file(self):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[0]
        self.ws.title = "Anki Cards"

    def write_cell(self, row, col, val):
        try:
            self.ws.cell(row=row, column=col).value = val
        except Exception as e:
            raise CannotWriteValueError(self.path, row, col, val) from e

    def write(self, notes, models, col_width):
        ws = self.ws
        first_line = []
        headers = []
        for model in models:
            first_line.append(model["name"])
            hd = [model["id"]]
            hd += model["flds"]
            headers.append(hd)

        # write headers
        for n in range(len(first_line)):
            self.write_cell(1, n + 1, first_line[n])
        for n in range(len(headers)):
            for m in range(len(headers[n])):
                self.write_cell(n + 2, m + 1, headers[n][m])

        # write notes
        crow = len(headers) + 1  # 1 row before first row of note rows
        for note in notes:
            crow += 1
            model = note.model()
            val_row = [None] * len(model["flds"])
            thismodel = None
            for mdl in models:
                if mdl["name"] == model["name"]:
                    thismodel = mdl
                    break
            if not thismodel:
                raise ModelNameDoesNotExistError(model["name"])

            self.write_cell(crow, 1, str(thismodel["id"]))
            for n in range(len(model["flds"])):
                for m in range(len(thismodel["flds"])):
                    if thismodel["flds"][m] == model["flds"][n]["name"]:
                        val_row[m] = note.fields[n]
                        break
            for n in range(len(val_row)):
                self.write_cell(crow, n + 2, str(val_row[n]))
            self.write_cell(crow, len(model["flds"]) + 2, note.id)

        for x in range(len(col_width)):
            cl = ws.cell(row=1, column=x + 1)
            if cl:
                col = cl.column_letter
                ws.column_dimensions[col].width = col_width[x]

    def save(self):
        dir = os.path.dirname(self.path)
        if not os.path.exists(dir):
            os.makedirs(dir)
        self.wb.save(filename=self.path)
